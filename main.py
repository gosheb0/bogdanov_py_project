import csv
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


class Vacancy:
    currency_to_rub = {
        "AZN": 35.68, "BYR": 23.91,
        "EUR": 59.90, "GEL": 21.74,
        "KGS": 0.76, "KZT": 0.13,
        "RUR": 1, "UAH": 1.64,
        "USD": 60.66, "UZS": 0.0055,
    }

    def __init__(self, val):
        self.name = val['name']
        self.salary_from = int(float(val['salary_from']))
        self.salary_to = int(float(val['salary_to']))
        self.salary_currency = val['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = val['area_name']
        self.year = int(val['published_at'][:4])


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    @staticmethod
    def increment(dic, k, score):
        if k not in dic:
            dic[k] = score
        else:
            dic[k] += score

    @staticmethod
    def average(value):
        new_dict = {}
        for k, v in value.items():
            new_dict[k] = int(sum(v) / len(v))
        return new_dict

    def csv_reader(self):
        with open(self.file_name, "r", encoding='utf-8-sig') as csv_file:
            reader = csv.reader(csv_file)
            header = next(reader)
            for row in reader:
                if '' not in row and len(row) == len(header):
                    yield dict(zip(header, row))

    def find_stats(self):
        count = 0

        salary = {}
        salary_vac = {}
        salary_city = {}

        for vac in self.csv_reader():
            vacancy = Vacancy(vac)
            self.increment(salary,
                           vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) == -1:
                pass
            else:
                self.increment(salary_vac, vacancy.year, [vacancy.salary_average])
            self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])

            count += 1

        vacancies_num = dict([(key, len(value)) for key, value in salary.items()])
        vacancies_num_name = dict([(key, len(value)) for key, value in salary_vac.items()])

        if salary_vac:
            pass
        else:
            salary_vac = dict([(key, [0]) for key, value in salary.items()])
            vacancies_num_name = dict([(key, 0) for key, value in vacancies_num.items()])
        stats, stats_2, stats_3, stats_5 = self.forming_stats(count, salary, salary_city, salary_vac)

        return stats, vacancies_num, stats_2, vacancies_num_name, stats_3, stats_5

    def forming_stats(self, count, salary, salary_city, salary_vac):
        stats = self.average(salary)
        stats_2 = self.average(salary_vac)
        stats_3 = self.average(salary_city)
        stats_4 = {}
        for year, salaries in salary_city.items():
            stats_4[year] = round(len(salaries) / count, 4)
        stats_4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value)
                                                        for key, value in stats_4.items()]))
        stats_4.sort(key=lambda a: a[-1], reverse=True)
        stats_5 = stats_4.copy()
        stats_4 = dict(stats_4)
        stats_3 = list(filter(lambda a: a[0] in list(stats_4.keys()),
                              [(key, value) for key, value in stats_3.items()]))
        stats_3.sort(key=lambda a: a[-1], reverse=True)
        stats_3 = dict(stats_3[:10])
        stats_5 = dict(stats_5[:10])
        return stats, stats_2, stats_3, stats_5

    @staticmethod
    def printing(stats1, stats2, stats3, stats4, stats5, stats6):
        print('Динамика уровня зарплат по годам: {0}'.format(stats1))
        print('Динамика количества вакансий по годам: {0}'.format(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stats4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stats5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stats6))


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        stats_1, stats_2, stats_3, stats_4, stats_5, stats_6 = dataset.find_stats()
        dataset.printing(stats_1, stats_2, stats_3, stats_4, stats_5, stats_6)

        report = Report(self.vacancy_name, stats_1, stats_2, stats_3, stats_4, stats_5, stats_6)
        report.generate_excel()
        report.save('report.xlsx')
        report.generating_graph()


def widths_column(data, l2):
    column_widths = []
    for row in data:
        for i, cell in enumerate(row):
            cell = str(cell)
            if len(column_widths) <= i:
                column_widths += [len(cell)]
            else:
                if len(cell) <= column_widths[i]:
                    continue
                column_widths[i] = len(cell)
    for i, column_width in enumerate(column_widths, 1):
        l2.column_dimensions[get_column_letter(i)].width = column_width + 2


class Report:
    def __init__(self, vacancy_name, stats_1, stats_2, stats_3, stats_4, stats_5, stats_6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats_1 = stats_1
        self.stats_2 = stats_2
        self.stats3 = stats_3
        self.stats_4 = stats_4
        self.stats_5 = stats_5
        self.stats_6 = stats_6

    def generate_excel(self):
        l1 = self.wb.active
        l1.title = 'Статистика по годам'
        l1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий',
                   'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats_1.keys():
            l1.append([year, self.stats_1[year], self.stats3[year], self.stats_2[year], self.stats_4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) <= i:
                    column_widths += [len(cell)]
                else:
                    if len(cell) <= column_widths[i]:
                        continue
                    column_widths[i] = len(cell)

        for i, column_width in enumerate(column_widths, 1):
            l1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city_1, value_1), (city_2, value_2) in zip(self.stats_5.items(), self.stats_6.items()):
            data.append([city_1, value_1, '', city_2, value_2])
        l2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            l2.append(row)
        widths_column(data, l2)
        self.font_bold(l1, l2)
        self.thin(data, l1, l2)

    def thin(self, data, l1, l2):
        thin = Side(border_style='thin', color='00000000')
        for row in range(len(data)):
            for v in 'ABDE':
                l2[v + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)
        for row, _ in enumerate(self.stats_1):
            for v in 'ABCDE':
                l1[v + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def font_bold(self, l1, l2):
        font_bold = Font(bold=True)
        for v in 'ABCDE':
            l1[v + '1'].font = font_bold
            l2[v + '1'].font = font_bold
        for i, _ in enumerate(self.stats_5):
            l2['E' + str(i + 2)].number_format = '0.00%'

    def generating_graph(self):
        fig, ((ax_1, ax_2), (ax_3, ax_4)) = plt.subplots(nrows=2, ncols=2)

        bar_1 = ax_1.bar(np.array(list(self.stats_1.keys())) - 0.4, self.stats_1.values(), width=0.4)
        bar_2 = ax_1.bar(np.array(list(self.stats_1.keys())), self.stats3.values(), width=0.4)
        self.forming_ax_1(ax_1, bar_1, bar_2)
        ax_2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar_1 = ax_2.bar(np.array(list(self.stats_2.keys())) - 0.4, self.stats_2.values(), width=0.4)
        bar_2 = ax_2.bar(np.array(list(self.stats_2.keys())), self.stats_4.values(), width=0.4)
        self.forming_ax_2(ax_2, bar_1, bar_2)
        self.forming_ax_3(ax_3)
        ax_4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([v for v in self.stats_6.values()])
        ax_4.pie(list(self.stats_6.values()) + [other], labels=list(self.stats_6.keys()) + ['Другие'],
                 textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    def forming_ax_3(self, ax_3):
        ax_3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax_3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n')
                        for a in reversed(list(self.stats_5.keys()))]),
                  list(reversed(list(self.stats_5.values()))), color='blue', height=0.5, align='center')
        ax_3.yaxis.set_tick_params(labelsize=6)
        ax_3.xaxis.set_tick_params(labelsize=8)
        ax_3.grid(axis='x')

    def forming_ax_2(self, ax_2, bar_1, bar_2):
        ax_2.legend((bar_1[0], bar_2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                    prop={'size': 8})
        ax_2.set_xticks(np.array(list(self.stats_2.keys())) - 0.2, list(self.stats_2.keys()), rotation=90)
        ax_2.grid(axis='y')
        ax_2.xaxis.set_tick_params(labelsize=8)
        ax_2.yaxis.set_tick_params(labelsize=8)

    def forming_ax_1(self, ax_1, bar_1, bar_2):
        ax_1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax_1.grid(axis='y')
        ax_1.legend((bar_1[0], bar_2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        ax_1.set_xticks(np.array(list(self.stats_1.keys())) - 0.2, list(self.stats_1.keys()), rotation=90)
        ax_1.xaxis.set_tick_params(labelsize=8)
        ax_1.yaxis.set_tick_params(labelsize=8)

    def save(self, filename):
        self.wb.save(filename=filename)


if __name__ == '__main__':
    InputConnect()

# Введите название файла: vacancies_by_year.csv
# Введите название профессии: Аналитик
