import csv
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


class Vacancy:
    """ Класс для представления вакансий

        Attributes:
            self.name (str): Название вакансии
            self.salary_from (str or int or float): Нижняя граница вилки оклада
            self.salary_to (str or int or float): Верхняя граница вилки оклада
            self.salary_currency (str): Валюта оклада
            self.salary_average : Средний оклад
            self.area_name (str):  Город публикиции вакансии
            self.year (int):  Год публикации вакансии
        """
    currency_to_rub = {
        "AZN": 35.68, "BYR": 23.91,
        "EUR": 59.90, "GEL": 21.74,
        "KGS": 0.76, "KZT": 0.13,
        "RUR": 1, "UAH": 1.64,
        "USD": 60.66, "UZS": 0.0055,
    }

    def __init__(self, val):
        self.name = val['name']
        self.salary_from = int(float(val['salary_from']))
        self.salary_to = int(float(val['salary_to']))
        self.salary_currency = val['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = val['area_name']
        self.year = int(val['published_at'][:4])


def increment(dic, k, score):
    if k not in dic:
        dic[k] = score
    else:
        dic[k] += score


class DataSet:
    def __init__(self, file_name, vac_name):
        self.file_name = file_name
        self.vacancy_name = vac_name

    @staticmethod
    def average(value):
        """
                Формирует среднее значение
                """
        new_dict = {}
        for k, v in value.items():
            new_dict[k] = int(sum(v) / len(v))
        return new_dict

    @property
    def csv_reader(self):
        """
                       Считывает файл, форммирует словарь
               """
        with open(self.file_name, "r", encoding='utf-8-sig') as csv_file:
            reader = csv.reader(csv_file)
            header = next(reader)
            for row in reader:
                if '' not in row and len(row) == len(header):
                    yield dict(zip(header, row))

    @property
    def find_stats(self):
        count = 0

        salary_vac = {}
        salary = {}
        salary_city = {}

        for vacancy_dictionary in self.csv_reader:
            vac = Vacancy(vacancy_dictionary)
            increment(salary, vac.year, [vac.salary_average])
            if vac.name.find(self.vacancy_name) == -1:
                pass
            else:
                increment(salary_vac, vac.year, [vac.salary_average])
            increment(salary_city, vac.area_name, [vac.salary_average])
            count += 1

        vacancies_num = dict([(key, len(value)) for key, value in salary.items()])
        vacancies_num_name = dict([(key, len(value)) for key, value in salary_vac.items()])

        if not salary_vac:
            salary_vac = dict([(key, [0]) for key, value in salary.items()])
            vacancies_num_name = dict([(key, 0) for key, value in vacancies_num.items()])

        stats, stats_2, stats_3, stats_5 = self.forming_stats(count, salary, salary_city, salary_vac)

        return stats, vacancies_num, stats_2, vacancies_num_name, stats_3, stats_5

    def forming_stats(self, count, salary, salary_city, salary_vac):
        stats = self.average(salary)
        stats_2 = self.average(salary_vac)
        stats_3 = self.average(salary_city)
        stats_4 = {}
        for year, sal in salary_city.items():
            stats_4[year] = round(len(sal) / count, 4)
        stats_4 = list(filter(lambda a: a[-1] >= 0.01,
                              [(key, value) for key, value in stats_4.items()]))
        stats_4.sort(key=lambda a: a[-1], reverse=True)
        stats_5 = stats_4.copy()
        stats_4 = dict(stats_4)
        stats_3 = list(filter(lambda a: a[0] in list(stats_4.keys()),
                              [(key, value) for key, value in stats_3.items()]))
        stats_3.sort(key=lambda a: a[-1], reverse=True)
        stats_3 = dict(stats_3[:10])
        stats_5 = dict(stats_5[:10])
        return stats, stats_2, stats_3, stats_5

    @staticmethod
    def print_statistic(stats1, stats2, stats3, stats4, stats5, stats6):
        print('Динамика уровня зарплат по годам: {0}'.format(stats1))
        print('Динамика количества вакансий по годам: {0}'.format(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stats4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stats5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stats6))


class InputConnect:
    """ Класс для ввода данных"""
    def __init__(self) -> object:
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        stats1, stats2, stats3, \
        stats4, stats5, stats6 = dataset.find_stats
        dataset.print_statistic(stats1, stats2,
                                stats3, stats4, stats5, stats6)

        report = Report(self.vacancy_name, stats1, stats2,
                        stats3, stats4, stats5, stats6)
        report.to_excel()


def widths_of_columns(data, ws2):
    column_widths = []
    for row in data:
        for i, v in enumerate(row):
            v = str(v)
            if len(column_widths) <= i:
                column_widths += [len(v)]
            else:
                if len(v) <= column_widths[i]:
                    continue
                column_widths[i] = len(v)
    for i, column_width in enumerate(column_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = column_width + 2


class Report:
    """ Класс для формирования отчета """
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats_1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats_5 = stats5
        self.stats_6 = stats6

    def to_excel(self):
        """ Формирования excel документа """
        l1 = self.wb.active
        l1.title = 'Статистика по годам'
        l1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name,
                   'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats_1.keys():
            l1.append([year, self.stats_1[year], self.stats3[year],
                       self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name,
                 ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, v in enumerate(row):
                if len(column_widths) <= i:
                    column_widths += [len(v)]
                else:
                    if len(v) <= column_widths[i]:
                        continue
                    column_widths[i] = len(v)

        for i, column_width in enumerate(column_widths, 1):
            l1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city_1, value_1), (city_2, value_2) in zip(self.stats_5.items(), self.stats_6.items()):
            data.append([city_1, value_1, '', city_2, value_2])
        l2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            l2.append(row)
        widths_of_columns(data, l2)
        self.font_of_bold(l1, l2)
        self.params_of_ex(data, l1, l2)
        self.wb.save('report.xlsx')

    def params_of_ex(self, data, l1, l2):
        thin = Side(border_style='thin', color='00000000')
        r = len(data)
        for row in range(r):
            for v in 'ABDE':
                l2[v + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)
        self.stats_1[1] = 1
        for row, _ in enumerate(self.stats_1):
            for v in 'ABCDE':
                l1[v + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def font_of_bold(self, l1, l2):
        """ Жирность шрифта """
        font_bold = Font(bold=True)
        for v in 'ABCDE':
            l1[v + '1'].font = font_bold
            l2[v + '1'].font = font_bold
        for i, _ in enumerate(self.stats_5):
            l2['E' + str(i + 2)].number_format = '0.00%'


if __name__ == '__main__':
    InputConnect()

# Введите название файла: vacancies_by_year.csv
# Введите название профессии: Аналитик
